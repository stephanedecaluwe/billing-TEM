# -*- coding: utf-8 -*-
# conversion synthèse ADV .xlsx => fichier de commande SAP .csv
# nov. 2020 
# librairies openpyxl remplacant xlrd deprecated

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys
import os

#PO_number = sys.argv[1]
#periode_long = sys.argv[2]

#PO_number ="TELECH TEM TRIM Q2.2021"
#PO_number ="TELECH TEM MOIS 04.2021" 
#periode_long = "2ieme trimestre 2021"
#periode_long = "avril 2021"



type = "local"


if type == "local":
   racine=r"c:\users\sdecaluwe\desktop\TEM_CONNECT local"
   repertoire_syntheses=racine+"\output\\"
   #path_lut = r"\\coprfil.usr.ingenico.loc\Public\ProfessionalServices_France\TEM\facturation TEM\TEM_CONNECT\data\LUT.xlsx"
   path_lut = r"c:\users\sdecaluwe\desktop\TEM_CONNECT local\data\LUT.xlsx"
   path_concat = racine+"\output\concatentation.csv"
   #repertoire_SAP = r"c:\users\sdecaluwe\desktop\TEM_CONNECT local\SAP\\"
   
else:
   repertoire_syntheses=r"\\Frprfil\data\FranceTeam\Marketing&Communication\14. Projets\TEM\facturation TEM\TEM_CONNECT\output\\"
   path_lut = r"\\Frprfil\data\FranceTeam\Marketing&Communication\14. Projets\TEM\facturation TEM\TEM_CONNECT\data\LUT.xlsx"
   path_concat = r"\\Frprfil\data\FranceTeam\Marketing&Communication\14. Projets\TEM\facturation TEM\TEM_CONNECT\SAP\concatentation.csv"
   repertoire_SAP = r"\\Frprfil\data\FranceTeam\Marketing&Communication\14. Projets\TEM\facturation TEM\TEM_CONNECT\SAP\\"


# PO number (purchase order). Numero de commande SAP, une commande SAP est l'injection de donnees dans SAP, c'est une opération technique.
# ne pas confondre avec une commande au sens commercial, document contractuel.
# ainsi, le numero de commande est pour la facturation TEM un simple mnemonique.
# "TELECH TEM TRIMESTRE" ou "TELECH TEM MENSUEL" sont les deux valeurs possibles
# ce champ est retrouve dans le fichier zlivfac extrait de SAP, et permet de trier les parties mensuelles et trimestrielles d'un meme client
# PO number et PO date (au mois pres) ne doivent jamais avoir ete utilises, sinon le message
# "commande deja existante" est produit
#slot="MOIS 12.2020"
#periode_long="decembre 2020"
#periode="TELECH TEM "+slot ;### le format de la zone N° Commande devra évaluer au fil des facturations 
# avec une partie commune et fixe « TELECH TEM TRIM Qx.aaaa » et « TELECH TEM MOIS mm.aaaa »
#organisation="8101" ou "7140" selon le code SAP client ## lecture dans la LUT ci-dessous

division="8120" #delivery plant
poste_tarif_nul="ZSEF" # colonne i 
poste_tarif_nonnul="ZSEV"

#repertoire ou sont stockes les syntheses ADV (1 ou plusieurs)
#repertoire_syntheses = r"\\Frprfil\data\FranceTeam\Marketing&Communication\14. Projets\TEM\facturation TEM\TEM_CONNECT\output\\"
# codes produit SAP qui sont ignores.

liste_exclusion = [ "T_AC-PRI-COM" ,"G_AC-CONNECT" , "G_AM-CONNECT" , "G3_AC-CONNECT" , "G3_AM-CONNECT" , "WP_AC-CONNECT" , "WP_AM-CONNECT"  ]

# codes client SAP qui sont retenus.

liste_exclusion_client = [ "80005337","80005036","80000671","80008251","80005269","SOLAM", "EDENRED", "SANTE MANUEL SEPHIRA", "SANTE Manuel", "80001289" ]

###= BP MED, CAPSYS OSB, Euro Information, EIS, ajout du 29/01 : SOLAM*, EDENRED*, ATA (automat...)
#	80005337	BP MEDITERRANEE	
#	80005036	CAPSYS	
#	80000671	OCEANIENNE DE SERVICES BANCAIRES	
#	80008251	EURO INFORMATION	
#	80005269	EURO INFORMATION SERVICES	
#   80009303	AUTOMATISME ET TECHNIQUES AVANCEES 
#   80001289    AFONE
#dans la LUT sont en "999"
#*SOLAM, *EDENRED, "Mig AUTO CB 5 5 et CRVPR1"
# les "999" sont exclus

## backup.liste_exclusion_client = [ "80005337","80005036","80000671","80008251","80005269","80012825", "EDENRED" , "SOLAM"] ###= BP MED, CAPSYS OSB, Euro Information, EIS.
# rattrapage si mauvaise syntaxe

dictionnaire_substitution = {}


# ouvrir le fichier LUT
#wb = xlrd.open_workbook(r"\\Frprfil\data\FranceTeam\Marketing&Communication\14. Projets\TEM\facturation TEM\TEM_CONNECT\data\LUT.xlsx")   

wb = Workbook()

wb = load_workbook(filename = path_lut)

# aller dans mapping retail TB

ws = wb["mapping retail TB"]

dico_entite={}

# remplir le dico entite : { code_sap client : sales_org 8101 ou 7140 dans l'onglet mapping

for row in range(ws.max_row):
   #print(row+1)
   code_client_SAP= str(ws.cell(row+1,2).value)
   entite=str(ws.cell(row+1,1).value)
   dico_entite[code_client_SAP]=entite

wb = Workbook()
wb = load_workbook(filename = path_lut)
ws = wb["Feuil1"]
tarif0=[]
for row in range(ws.max_row):
   if str(ws.cell(row+1,6).value)=='0' and str(ws.cell(row+1,2).value)!='non facturable':
      tarif0.append(str(ws.cell(row+1,2).value))

tarif0.append("TEM_AUTO_CHAR0_4")
tarif0.append("TEM_AUTO_INIT0_4")
   
print(tarif0)
 
   
# parcourir les champs, remplir le dico en excluant titre et champs vides.
#  if str(ws.cell(row,1))[-2:]==".0":
#   if str(ws.cell(row+1,1).value)=="0":
#      code_client_SAP= str(ws.cell(row,1).value)
#   else:
#      code_client_SAP= str(ws.cell(row+1,1).value)
#   entite=str(ws.cell(row+1,1).value)
#   if code_client_SAP !="Sold-to par" and code_client_SAP !="":
#      dico_entite[code_client_SAP]=entite

#print(dico_entite)


# effacement de SAP/*.csv
for i in os.listdir(repertoire_syntheses):
   if i.endswith(".csv"):
      os.remove(repertoire_syntheses+i)
      
fichier_total = open(path_concat,'a')   

# pour tous les fichiers du repertoire de sortie de tem connect type synthèse ADV*.xlsx
for i in os.listdir(repertoire_syntheses):
    # si le fichier s'appelle synthese adv*.xlsx
    if i.startswith("synthèse ADV") and (i.endswith(".xlsx") or i.endswith(".XLSX")):
       print("traitemenent de ........ "+i)
       wb = load_workbook(filename = repertoire_syntheses+i)
       # aller dans le premier onglet (numero 0)
       ws = wb.worksheets[0]
       # la date de commande SAP (PO) est prise dans le champ "à" de la synthèse ADV.xlsx (ex : du 01/07/2020 à 30/09/2020)
       PO_date=str(ws.cell(2,13).value)
       PO_datedebut=str(ws.cell(2,12).value)
       annee = PO_date[-4:]
       mois1 = PO_datedebut[3:5]
       mois2 = PO_date[3:5]
       mois = mois2
       intervalle = int(mois2)-int(mois1)
       print(intervalle)
       if intervalle == 0:
          PO_number="TELECH TEM MOIS "+mois+"."+annee
          if mois=="01":
             periode_long="janvier "+annee
          if mois=="02":
             periode_long="fevrier "+annee
          if mois=="03":
             periode_long="mars "+annee             
          if mois=="04":
             periode_long="avril "+annee
          if mois=="05":
             periode_long="mai "+annee             
          if mois=="06":
             periode_long="juin "+annee             
          if mois=="07":
             periode_long="juillet "+annee                
          if mois=="08":
             periode_long="aout "+annee    
          if mois=="09":
             periode_long="septembre "+annee    
          if mois=="10":
             periode_long="octobre "+annee    
          if mois=="11":
             periode_long="novembre "+annee    
          if mois=="12":
             periode_long="decembre "+annee
       if intervalle == 2:
          if mois=="03":
             PO_number="TELECH TEM TRIM Q1 "+annee
             periode_long="1er trimestre "+annee
          if mois=="06":
             PO_number="TELECH TEM TRIM Q2 "+annee
             periode_long="2eme trimestre "+annee
          if mois=="09":
             PO_number="TELECH TEM TRIM Q3 "+annee
             periode_long="3eme trimestre "+annee
          if mois=="12":
             PO_number="TELECH TEM TRIM Q4 "+annee
             periode_long="4eme trimestre "+annee   
       # le code client SAP est pris dans la colonne A (indice 0)
       #print(ws)
       #print(ws.cell(2,1).value)
       code_client= str(ws.cell(2,1).value)
       code_client_precedent=42
       organisation = dico_entite[code_client]
       # le nom du client, dans la colonne B (indice 1)
       nom_client=str(ws.cell(2,2).value)
       # assemblage des chaines selon le format final : nom fichier, ORG, HEADER.
       nom_fichier=repertoire_syntheses+nom_client+" "+PO_date.replace("/","-")+".csv"
       fichier  = open(nom_fichier,'a')
       for row in range(ws.max_row):
          # exclusion des noms de champs
          
          if ws.cell(row+1,1).value != "" and ws.cell(row+1,1).value !="code SAP" and ws.cell(row+1,1).value !="Code SAP" and ws.cell(row+1,1).value !="code SAP client":
             # code SAP produit, quantite;
             code_client= str(ws.cell(row+1,1).value)
             #print(code_client)
             if code_client!=code_client_precedent and code_client not in liste_exclusion_client:
                # un nouveau client est traité : ORG, HEADER et TEXTH sont mis en en-tête
                #print(code_client)
                organisation = dico_entite[code_client]
                fichier.write("ORG;ZSCE;"+str(organisation)+";Z1;SE;;;;\n")
                fichier.write("HEADER;"+PO_number+";"+PO_date+";;"+code_client+";"+code_client+";;;\n")
                fichier.write("TEXTH;0011;FR;Facturation des telechargements ESTATE MANAGER;;;;;\n")
                fichier.write("TEXTH;0011;FR;Periode :  "+periode_long+";;;;;\n")
                fichier.write("TEXTH;0011;FR;Merci d'adresser votre demande de justificatif a :;;;;;\n")
                fichier.write("TEXTH;0011;FR;adv-telechargements@ingenico.com;;;;;\n")
                fichier_total.write("ORG;ZSCE;"+str(organisation)+";Z1;SE;;;;\n")
                fichier_total.write("HEADER;"+PO_number+";"+PO_date+";;"+code_client+";"+code_client+";;;\n")
                fichier_total.write("TEXTH;0011;FR;Facturation des telechargements ESTATE MANAGER;;;;;\n")
                fichier_total.write("TEXTH;0011;FR;Periode :  "+periode_long+";;;;;\n")
                fichier_total.write("TEXTH;0011;FR;Merci d'adresser votre demande de justificatif a :;;;;;\n")
                fichier_total.write("TEXTH;0011;FR;adv-telechargements@ingenico.com;;;;;\n")
                code_client_precedent=code_client
             code_produit = str(ws.cell(row+1,7).value)
             quantite = str(ws.cell(row+1,9).value)
             # exclusion et substitution du code produit le cas echeant
             
             # écriture d'une ligne correspondant à un produit, quantité, client.
             # codes SAP exclus : non reconnus par SAP car à créer.
             # Les codes SAP à tarif nul sont traités avec le poste_tarif_nul (ZSEN, ZSEF ...).
             # application d'une liste de substitution pour garantir le bon format des codes SAP (corrigé maintenant).
             if code_produit not in liste_exclusion and code_client not in liste_exclusion_client:
                if code_produit not in dictionnaire_substitution:
                   if code_produit not in tarif0:
                      fichier.write("ITEM;"+code_produit+";"+quantite+";;EUR;;"+division+";;"+poste_tarif_nonnul+"\n")
                      fichier_total.write("ITEM;"+code_produit+";"+quantite+";;EUR;;"+division+";;"+poste_tarif_nonnul+"\n")
                      #print("ITEM;"+code_produit+";"+quantite+";;EUR;;"+division+";;"+poste_tarif_nonnul+"\n")
                      # cas standard : tarif_nonnul
                   else:
                      fichier.write("ITEM;"+code_produit+";"+quantite+";;EUR;;"+division+";;"+poste_tarif_nul+"\n")
                      fichier_total.write("ITEM;"+code_produit+";"+quantite+";;EUR;;"+division+";;"+poste_tarif_nul+"\n")
                      #print("ITEM;"+code_produit+";"+quantite+";;EUR;;"+division+";;"+poste_tarif_nul+"\n")
                      #tarif0 non exclu, non substitué
                else:                 
                   fichier.write("ITEM;"+dictionnaire_substitution[code_produit]+";"+quantite+";;EUR;;"+division+";;\n")
                   fichier_total.write("ITEM;"+dictionnaire_substitution[code_produit]+";"+quantite+";;EUR;;"+division+";;\n")
                   #print("ITEM;"+dictionnaire_substitution[code_produit]+";"+quantite+";;EUR;;"+division+";;\n")
                   #tarif standard substitué
       # sauvegarde, fermeture
       fichier.close()
       
fichier_total.write("") 
   
fichier_total.close()

print("fichiers écrits")   
 
       
